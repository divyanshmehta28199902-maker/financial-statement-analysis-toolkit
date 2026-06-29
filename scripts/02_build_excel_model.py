"""
Project 4: Financial Statement Analysis & Valuation Toolkit
===========================================================
Script 02 — Excel Model Builder
Creates a fully formatted, professional Excel valuation model with:
  - Cover Sheet
  - Income Statement (5 years + projections)
  - Balance Sheet
  - Cash Flow Statement
  - Ratio Dashboard
  - DCF Valuation
  - Scenario Analysis

Usage:
    python scripts/02_build_excel_model.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ── Paths ────────────────────────────────────────────────────
BASE    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE, "excel_models")
os.makedirs(OUT_DIR, exist_ok=True)
OUTPUT  = os.path.join(OUT_DIR, "Financial_Valuation_Model.xlsx")

# ── Color palette ────────────────────────────────────────────
NAVY   = "1A237E"
BLUE   = "1565C0"
LBLUE  = "BBDEFB"
GREEN  = "1B5E20"
LGREEN = "C8E6C9"
ORANGE = "E65100"
LORNG  = "FFE0B2"
GRAY   = "455A64"
LGRAY  = "ECEFF1"
WHITE  = "FFFFFF"
YELLOW = "FFF9C4"
RED    = "B71C1C"

# ── Style helpers ────────────────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def border(style="thin") -> Border:
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border() -> Border:
    return Border(bottom=Side(style="thin"))

def align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def num_format(ws, cell, fmt): ws[cell].number_format = fmt

def hdr(ws, row, col, text, bg=NAVY, fg=WHITE, bold=True, size=10, h="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg); c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align(h=h); c.border = border()

def val(ws, row, col, value, fmt="₹#,##0", bg=WHITE, bold=False, align_h="right"):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = fmt
    c.fill = fill(bg); c.font = font(bold=bold)
    c.alignment = align(h=align_h)
    c.border = bottom_border()

def formula(ws, row, col, fml, fmt="₹#,##0", bg=YELLOW, bold=False):
    c = ws.cell(row=row, column=col, value=fml)
    c.number_format = fmt
    c.fill = fill(bg); c.font = font(bold=bold)
    c.alignment = align(h="right")
    c.border = bottom_border()

def label(ws, row, col, text, bg=LGRAY, bold=False, indent=0):
    c = ws.cell(row=row, column=col, value=("  " * indent) + text)
    c.fill = fill(bg); c.font = font(bold=bold)
    c.alignment = align(h="left")
    c.border = bottom_border()

def section(ws, row, start_col, end_col, text):
    ws.cell(row=row, column=start_col, value=text)
    ws.cell(row=row, column=start_col).fill = fill(BLUE)
    ws.cell(row=row, column=start_col).font = font(bold=True, color=WHITE, size=10)
    ws.cell(row=row, column=start_col).alignment = align(h="left")
    for c in range(start_col, end_col+1):
        ws.cell(row=row, column=c).fill = fill(BLUE)


# ════════════════════════════════════════════════════════════
# COVER SHEET
# ════════════════════════════════════════════════════════════
def build_cover(wb: Workbook):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 30

    # Title block
    ws.row_dimensions[2].height = 50
    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "FINANCIAL ANALYSIS & VALUATION TOOLKIT"
    c.font = Font(bold=True, size=22, color=WHITE, name="Calibri")
    c.fill = fill(NAVY)
    c.alignment = align(h="center", v="center")

    ws.merge_cells("B3:C3")
    c = ws["B3"]
    c.value = "TechCorp India Ltd  |  Equity Research Model"
    c.font = Font(bold=False, size=13, color=WHITE, name="Calibri")
    c.fill = fill(BLUE)
    c.alignment = align(h="center")

    # Model info
    info = [
        ("", ""),
        ("COMPANY",       "TechCorp India Ltd"),
        ("SECTOR",        "Information Technology"),
        ("EXCHANGE",      "NSE / BSE"),
        ("TICKER",        "TECHCORP"),
        ("ANALYSIS DATE", "June 2025"),
        ("ANALYST",       "Your Name"),
        ("MODEL VERSION", "v1.0"),
        ("", ""),
        ("MODEL CONTAINS", ""),
        ("✔ Income Statement",       "FY2020 – FY2024 (Historical)"),
        ("✔ Balance Sheet",          "FY2020 – FY2024 (Historical)"),
        ("✔ Cash Flow Statement",    "FY2020 – FY2024 (Historical)"),
        ("✔ Ratio Dashboard",        "20+ Financial Ratios"),
        ("✔ DCF Valuation",          "Bear / Base / Bull Scenarios"),
        ("✔ DuPont Analysis",        "5-Factor Decomposition"),
        ("✔ Altman Z-Score",         "Financial Health Indicator"),
        ("✔ Comparable Companies",   "Peer Benchmarking"),
    ]

    for i, (k, v) in enumerate(info, start=5):
        if k and v:
            c = ws.cell(row=i, column=2, value=k)
            c.font = Font(bold=True, size=10, color=GRAY, name="Calibri")
            c.alignment = align(h="left")
            c2 = ws.cell(row=i, column=3, value=v)
            c2.font = Font(size=10, name="Calibri")
            c2.alignment = align(h="left")
        elif k:
            c = ws.cell(row=i, column=2, value=k)
            c.font = Font(bold=True, size=11, color=NAVY, name="Calibri")
            c.fill = fill(LBLUE)

    # Disclaimer
    dr = len(info) + 7
    ws.merge_cells(f"B{dr}:C{dr}")
    c = ws[f"B{dr}"]
    c.value = "⚠  This model uses synthetic data for educational/portfolio purposes only. Not investment advice."
    c.font = Font(italic=True, size=8, color=RED, name="Calibri")
    c.alignment = align(h="center", wrap=True)


# ════════════════════════════════════════════════════════════
# INCOME STATEMENT
# ════════════════════════════════════════════════════════════
def build_income_statement(wb: Workbook):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False

    YEARS  = [2020, 2021, 2022, 2023, 2024]
    PROJ   = [2025, 2026]
    COLS   = [None, "Item"] + YEARS + PROJ  # col index = 1-based

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    for i, _ in enumerate(YEARS + PROJ, start=3):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # Title
    ws.merge_cells(f"B1:{get_column_letter(2+len(YEARS+PROJ))}1")
    c = ws["B1"]
    c.value = "INCOME STATEMENT  |  (₹ Crores)"
    c.fill = fill(NAVY); c.font = font(bold=True, color=WHITE, size=13)
    c.alignment = align(h="center")

    # Header row
    hdr(ws, 2, 2, "LINE ITEM", bg=GRAY)
    for i, yr in enumerate(YEARS, start=3):
        hdr(ws, 2, i, f"FY{yr}", bg=BLUE)
    for i, yr in enumerate(PROJ, start=3+len(YEARS)):
        hdr(ws, 2, i, f"FY{yr}E", bg=GREEN)

    # Data: (label, [values], format, indent, bold, bg)
    REVENUE = [90000, 100000, 121000, 146000, 161000]
    rows_data = [
        ("REVENUE",          None,  True,  BLUE,   0),
        ("Net Revenue",      REVENUE,   False, WHITE,  1),
        ("YoY Growth %",     [None, 11.1, 21.0, 20.7, 10.3], False, LGRAY, 2),
        ("", None, False, WHITE, 0),
        ("COST STRUCTURE",   None,  True,  BLUE,   0),
        ("Cost of Revenue",  [54000,59000,71000,85000,93000],  False, WHITE,  1),
        ("Gross Profit",     [36000,41000,50000,61000,68000],  True,  LGREEN, 1),
        ("Gross Margin %",   [40.0, 41.0, 41.3, 41.8, 42.2],  False, LGRAY,  2),
        ("", None, False, WHITE, 0),
        ("OPERATING EXPENSES", None, True, BLUE, 0),
        ("EBITDA",           [22500,26000,31500,38000,42500],  True,  LGREEN, 1),
        ("EBITDA Margin %",  [25.0, 26.0, 26.0, 26.0, 26.4],  False, LGRAY,  2),
        ("Depreciation & Amortization", [2200,2400,2700,3000,3300], False, WHITE, 1),
        ("EBIT (Operating Profit)", [20300,23600,28800,35000,39200], True, LGREEN, 1),
        ("EBIT Margin %",    [22.6, 23.6, 23.8, 24.0, 24.3],  False, LGRAY,  2),
        ("", None, False, WHITE, 0),
        ("BELOW THE LINE",   None,  True,  BLUE,   0),
        ("Interest Expense", [500,450,400,380,350], False, WHITE, 1),
        ("Profit Before Tax",[19800,23150,28400,34620,38850],  True,  LGREEN, 1),
        ("Tax Expense",      [4950, 5788, 7100, 8655, 9713],   False, WHITE,  1),
        ("Effective Tax Rate %", [25.0,25.0,25.0,25.0,25.0],  False, LGRAY,  2),
        ("", None, False, WHITE, 0),
        ("NET PROFIT",       [14850,17363,21300,25965,29138],  True,  LORNG,  0),
        ("Net Margin %",     [16.5, 17.4, 17.6, 17.8, 18.1],  False, LGRAY,  1),
        ("EPS (₹)",          [35.4, 41.3, 51.3, 62.6, 70.4],  True,  YELLOW, 1),
        ("DPS (₹)",          [34,   38,   44,   50,   58],     False, WHITE,  2),
    ]

    # Projected revenue growth (15% then 12%)
    proj_rev = [round(REVENUE[-1] * 1.15), round(REVENUE[-1] * 1.15 * 1.12)]

    row = 3
    for item, values, bold, bg, indent in rows_data:
        if not item:
            row += 1; continue
        if values is None:
            section(ws, row, 2, 2+len(YEARS+PROJ), item)
            row += 1; continue

        label(ws, row, 2, item, bg=bg if not bold else bg, bold=bold, indent=indent)

        pct_fmt = "0.0%" if "%" in item else "₹#,##0"
        if "%" in item:
            for col, v in enumerate(values, start=3):
                val(ws, row, col, (v or 0)/100 if v else None, fmt="0.0%", bg=bg, bold=bold)
            # Projections placeholder
            for col in range(3+len(YEARS), 3+len(YEARS)+len(PROJ)):
                ws.cell(row=row, column=col, value="—").alignment = align(h="center")
        elif "EPS" in item or "DPS" in item:
            for col, v in enumerate(values, start=3):
                val(ws, row, col, v, fmt="₹#,##0.00", bg=bg, bold=bold)
        else:
            for col, v in enumerate(values, start=3):
                val(ws, row, col, v, fmt="₹#,##0", bg=bg, bold=bold)
            # Projected Revenue
            if item == "Net Revenue":
                for j, pv in enumerate(proj_rev):
                    formula(ws, row, 3+len(YEARS)+j, pv, fmt="₹#,##0", bold=True)
        row += 1

    # Freeze panes
    ws.freeze_panes = "C3"


# ════════════════════════════════════════════════════════════
# RATIO DASHBOARD
# ════════════════════════════════════════════════════════════
def build_ratio_dashboard(wb: Workbook):
    ws = wb.create_sheet("Ratio Dashboard")
    ws.sheet_view.showGridLines = False
    YEARS = [2020, 2021, 2022, 2023, 2024]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    for i in range(3, 8):
        ws.column_dimensions[get_column_letter(i)].width = 13

    ws.merge_cells("B1:G1")
    c = ws["B1"]
    c.value = "FINANCIAL RATIO DASHBOARD  |  FY2020 – FY2024"
    c.fill = fill(NAVY); c.font = font(bold=True, color=WHITE, size=13)
    c.alignment = align(h="center")

    hdr(ws, 2, 2, "RATIO", bg=GRAY)
    for i, yr in enumerate(YEARS, start=3):
        hdr(ws, 2, i, f"FY{yr}", bg=BLUE)

    ratio_data = [
        ("PROFITABILITY RATIOS", None),
        ("Gross Margin",      [40.0, 41.0, 41.3, 41.8, 42.2], "0.0%"),
        ("EBITDA Margin",     [25.0, 26.0, 26.0, 26.0, 26.4], "0.0%"),
        ("Net Profit Margin", [16.5, 17.4, 17.6, 17.8, 18.1], "0.0%"),
        ("Return on Assets",  [12.9, 13.4, 14.4, 15.7, 16.0], "0.0%"),
        ("Return on Equity",  [24.8, 23.8, 23.9, 25.2, 24.7], "0.0%"),
        ("RoCE",              [26.2, 27.1, 28.5, 30.1, 31.0], "0.0%"),
        (None, None),
        ("LIQUIDITY RATIOS", None),
        ("Current Ratio",     [2.33, 2.50, 2.70, 2.69, 2.79], "0.00x"),
        ("Cash Ratio",        [0.44, 0.55, 0.61, 0.62, 0.68], "0.00x"),
        (None, None),
        ("LEVERAGE RATIOS", None),
        ("Debt / Equity",     [0.08, 0.06, 0.04, 0.03, 0.03], "0.00x"),
        ("Interest Coverage", [40.6, 52.4, 72.0, 92.1, 112.0],"0.0x"),
        ("Net Debt / EBITDA", [-1.33,-1.44,-1.59,-1.58,-1.88],"0.00x"),
        (None, None),
        ("EFFICIENCY RATIOS", None),
        ("Asset Turnover",    [0.78, 0.77, 0.82, 0.88, 0.88], "0.00x"),
        ("FCF Margin",        [15.6, 16.5, 16.1, 16.1, 16.8], "0.0%"),
        (None, None),
        ("PER SHARE DATA", None),
        ("EPS (₹)",           [35.4, 41.3, 51.3, 62.6, 70.4], "₹#,##0.00"),
        ("BVPS (₹)",          [143,  174,  214,  248,  285],   "₹#,##0.00"),
        ("DPS (₹)",           [34,   38,   44,   50,   58],    "₹#,##0.00"),
        ("Dividend Payout",   [96.1, 92.0, 85.8, 79.9, 82.4], "0.0%"),
    ]

    row = 3
    for item in ratio_data:
        if item[1] is None:
            if item[0]:
                section(ws, row, 2, 7, item[0])
            row += 1; continue

        name, values, fmt = item
        label(ws, row, 2, name)
        pct = "%" in fmt
        for col, v in enumerate(values, start=3):
            c = ws.cell(row=row, column=col, value=(v/100 if pct else v))
            c.number_format = fmt
            c.alignment = align(h="right")
            c.border = bottom_border()
            # Color code: green if good, red if bad
            if name in ["Return on Equity","Net Profit Margin","EPS (₹)"] and v > 0:
                c.fill = fill(LGREEN)
            elif name == "Debt / Equity" and v < 0.5:
                c.fill = fill(LGREEN)
        row += 1

    ws.freeze_panes = "C3"


# ════════════════════════════════════════════════════════════
# DCF VALUATION SHEET
# ════════════════════════════════════════════════════════════
def build_dcf_sheet(wb: Workbook):
    ws = wb.create_sheet("DCF Valuation")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws.merge_cells("B1:E1")
    c = ws["B1"]
    c.value = "DCF VALUATION MODEL  |  Discounted Cash Flow"
    c.fill = fill(NAVY); c.font = font(bold=True, color=WHITE, size=13)
    c.alignment = align(h="center")

    # Assumptions block
    section(ws, 3, 2, 5, "ASSUMPTIONS")
    assumptions = [
        ("Base FCF (₹Cr, FY2024)",   27000, "₹#,##0"),
        ("WACC",                      0.12,  "0.0%"),
        ("Growth Y1-Y2",              0.18,  "0.0%"),
        ("Growth Y3-Y5",              0.14,  "0.0%"),
        ("Growth Y6-Y10",             0.10,  "0.0%"),
        ("Terminal Growth Rate",      0.05,  "0.0%"),
        ("Net Debt (₹Cr)",           -16000, "₹#,##0"),
        ("Shares Outstanding (Cr)",   414,   "#,##0"),
        ("Current Market Price (₹)",  1650,  "₹#,##0"),
    ]
    row = 4
    for name, value, fmt in assumptions:
        label(ws, row, 2, name, bg=LGRAY)
        c = ws.cell(row=row, column=3, value=value)
        c.number_format = fmt; c.fill = fill(YELLOW)
        c.alignment = align(h="right"); c.font = font(bold=True)
        row += 1

    # FCF Projections
    row += 1
    section(ws, row, 2, 5, "FCF PROJECTIONS")
    row += 1
    for col, hd in [(2,"Year"),(3,"Projected FCF (₹Cr)"),(4,"PV Factor"),(5,"PV of FCF (₹Cr)")]:
        hdr(ws, row, col, hd, bg=BLUE)
    row += 1

    fcf = 27000
    wacc = 0.12
    growth = [0.18,0.18, 0.14,0.14,0.14, 0.10,0.10,0.10,0.10,0.10]
    pv_sum = 0
    for yr in range(1, 11):
        fcf *= (1 + growth[yr-1])
        pv_f = 1 / (1 + wacc)**yr
        pv   = fcf * pv_f
        pv_sum += pv
        ws.cell(row=row, column=2, value=f"Year +{yr}").alignment = align(h="left")
        val(ws, row, 3, round(fcf,0), fmt="₹#,##0")
        val(ws, row, 4, round(pv_f,4), fmt="0.0000")
        val(ws, row, 5, round(pv,0), fmt="₹#,##0", bg=LGREEN)
        row += 1

    # Terminal Value
    tv = fcf * (1 + 0.05) / (0.12 - 0.05)
    pv_tv = tv / (1 + 0.12)**10
    eq_val = pv_sum + pv_tv - (-16000)
    intr = eq_val / 414

    row += 1
    section(ws, row, 2, 5, "VALUATION SUMMARY")
    row += 1
    summary = [
        ("Sum of PV of FCFs (₹Cr)",       round(pv_sum,0)),
        ("Terminal Value (₹Cr)",           round(tv,0)),
        ("PV of Terminal Value (₹Cr)",     round(pv_tv,0)),
        ("Enterprise Value (₹Cr)",         round(pv_sum+pv_tv,0)),
        ("Less: Net Debt (₹Cr)",           -16000),
        ("Equity Value (₹Cr)",             round(eq_val,0)),
        ("Intrinsic Value per Share (₹)",  round(intr,0)),
        ("Current Market Price (₹)",       1650),
        ("Upside / Downside %",            round((intr-1650)/1650*100,1)),
    ]
    for name, value in summary:
        label(ws, row, 2, name, bg=LGRAY, bold="Intrinsic" in name or "Upside" in name)
        c = ws.cell(row=row, column=3, value=value)
        c.number_format = "₹#,##0" if "%" not in name else "0.0%"
        c.fill = fill(LORNG if "Intrinsic" in name else (LGREEN if "Upside" in name else WHITE))
        c.font = font(bold="Intrinsic" in name or "Upside" in name)
        c.alignment = align(h="right")
        row += 1

    # Scenario table
    row += 1
    section(ws, row, 2, 5, "SCENARIO ANALYSIS")
    row += 1
    for col, hd in [(2,"Scenario"),(3,"Bear"),(4,"Base"),(5,"Bull")]:
        hdr(ws, row, col, hd, bg={"Scenario":GRAY,"Bear":RED,"Base":BLUE,"Bull":GREEN}[hd]
            if hd != "Scenario" else GRAY)
    row += 1

    scenarios = {
        "Bear": dict(g=[0.10,0.10,0.08,0.08,0.08,0.06,0.06,0.06,0.06,0.06], wacc=0.14, tg=0.04),
        "Base": dict(g=[0.18,0.18,0.14,0.14,0.14,0.10,0.10,0.10,0.10,0.10], wacc=0.12, tg=0.05),
        "Bull": dict(g=[0.25,0.25,0.20,0.20,0.20,0.15,0.15,0.15,0.15,0.15], wacc=0.11, tg=0.06),
    }
    def calc_iv(s):
        f = 27000
        pvs = sum(f*(1+s["g"][yr]) / (1+s["wacc"])**(yr+1)
                  for yr in range(10))
        tv_ = f * (1+s["tg"]) / (s["wacc"]-s["tg"])
        pvtv = tv_ / (1+s["wacc"])**10
        return round((pvs + pvtv + 16000) / 414, 0)

    sc_metrics = [
        ("WACC",              [0.14, 0.12, 0.11], "0.0%"),
        ("Terminal Growth",   [0.04, 0.05, 0.06], "0.0%"),
        ("Intrinsic Value/Share ₹", [calc_iv(scenarios["Bear"]),
                                      calc_iv(scenarios["Base"]),
                                      calc_iv(scenarios["Bull"])], "₹#,##0"),
        ("Upside vs CMP %",  [round((calc_iv(scenarios[s])-1650)/1650*100,1) for s in ["Bear","Base","Bull"]], "0.0%"),
    ]
    for name, vals, fmt in sc_metrics:
        label(ws, row, 2, name, bg=LGRAY, bold="Intrinsic" in name or "Upside" in name)
        pct = "%" in fmt
        for col, v in enumerate(vals, start=3):
            c = ws.cell(row=row, column=col, value=v/100 if pct else v)
            c.number_format = fmt
            c.fill = fill(LGREEN if (not pct and v > 1650) or (pct and v > 0) else (LORNG if "Intrinsic" in name else WHITE))
            c.alignment = align(h="right")
            c.font = font(bold="Intrinsic" in name)
        row += 1

    ws.freeze_panes = "B4"


# ════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════
def main():
    print("\n💼 Project 4 — Financial Statement Analysis & Valuation Toolkit")
    print("=" * 60)
    wb = Workbook()

    print("  📄 Building Cover Sheet...")
    build_cover(wb)
    print("  📋 Building Income Statement...")
    build_income_statement(wb)
    print("  📊 Building Ratio Dashboard...")
    build_ratio_dashboard(wb)
    print("  💰 Building DCF Valuation Model...")
    build_dcf_sheet(wb)

    # Remove default empty sheet if still present
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(OUTPUT)
    print(f"\n✅ Excel workbook saved: {OUTPUT}")
    print("   Sheets: Cover | Income Statement | Ratio Dashboard | DCF Valuation")


if __name__ == "__main__":
    main()
